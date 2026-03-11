from __future__ import annotations

import csv
import json
import time
from collections import Counter
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Callable
import unicodedata

import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.schemas.lite import LiteAnalyzeResponse, LitePreviewRow, LiteResultRow, LiteRunResponse, LiteRunSummary
from app.services.lite_status_mapper import LiteStatusResult, map_route_response
from app.services.settings_service import SettingsService
from app.services.sf_client import SFClient, SFClientCredentials, SFClientError
from app.utils.excel_safe import escape_excel_formula

MAX_LITE_FETCH_CONCURRENCY = 3
PARTIAL_MISSING_CODE = "SF_PARTIAL_MISSING"
PARTIAL_MISSING_REMARK = "SF batch response missing tracking number"

FIELD_ALIASES: dict[str, list[str]] = {
    # 고객사마다 컬럼명이 달라서 대표 별칭을 넓게 받아 자동 감지한다.
    "order_number": [
        "order_number",
        "order number",
        "order no",
        "order_no",
        "orderno",
        "ordernumber",
    ],
    "tracking_number": [
        "tracking_number",
        "tracking number",
        "tracking no",
        "tracking_no",
        "trackingnumber",
        "waybill",
        "waybill_no",
        "tracking",
        "mailno",
    ],
}

EXPORT_COLUMNS = [
    ("쇼핑몰오더번호", "order_number"),
    ("송장번호", "tracking_number"),
    ("송장상태", "status"),
    ("택배사최신상태코드", "sf_express_code"),
    ("택배사최신REMARK", "sf_express_remark"),
]

UNKNOWN_LOG_COLUMNS = [
    ("쇼핑몰오더번호", "order_number"),
    ("송장번호", "tracking_number"),
    ("송장상태", "status"),
    ("택배사최신상태코드", "sf_express_code"),
    ("택배사최신REMARK", "sf_express_remark"),
    ("최신이벤트일시", "last_event_time"),
    ("OP CODE", "op_code"),
    ("1차상태코드", "first_status_code"),
    ("2차상태코드", "secondary_status_code"),
    ("최신이벤트원문", "latest_event_raw"),
]

EXPORT_STATUS_LABELS = {
    "QUERY_UNAVAILABLE": "조회불가",
}

EXPORT_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
EXPORT_HEADER_FONT = Font(bold=True)
EXPORT_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


@dataclass
class LiteInputRow:
    order_number: str
    tracking_number: str | None


class LiteService:
    def __init__(self, session: Session, settings: Settings) -> None:
        self.session = session
        self.settings = settings
        self.settings_service = SettingsService(session, settings)

    async def analyze_upload(
        self,
        upload_file: UploadFile,
        mapping_override: dict[str, str | None] | None = None,
        sheet_name: str | None = None,
    ) -> LiteAnalyzeResponse:
        file_name, content = await self._read_upload(upload_file)
        prepared = self.prepare_content(
            file_name=file_name,
            content=content,
            mapping_override=mapping_override,
            sheet_name=sheet_name,
            validate_required_mapping=False,
        )
        return self._build_analyze_response(prepared)

    async def run_upload(
        self,
        upload_file: UploadFile,
        mapping_override: dict[str, str | None] | None = None,
        sheet_name: str | None = None,
        batch_size: int = 10,
        delay_seconds: float = 0.0,
        language: str = "0",
    ) -> LiteRunResponse:
        file_name, content = await self._read_upload(upload_file)
        prepared = self.prepare_content(
            file_name=file_name,
            content=content,
            mapping_override=mapping_override,
            sheet_name=sheet_name,
            validate_required_mapping=True,
        )
        return self.run_prepared(
            prepared=prepared,
            batch_size=batch_size,
            delay_seconds=delay_seconds,
            language=language,
        )

    async def export_upload(
        self,
        upload_file: UploadFile,
        mapping_override: dict[str, str | None] | None = None,
        sheet_name: str | None = None,
        file_format: str = "xlsx",
        batch_size: int = 10,
        delay_seconds: float = 0.0,
        language: str = "0",
    ) -> tuple[str, bytes, str]:
        file_name, content = await self._read_upload(upload_file)
        prepared = self.prepare_content(
            file_name=file_name,
            content=content,
            mapping_override=mapping_override,
            sheet_name=sheet_name,
            validate_required_mapping=True,
        )
        result = self.run_prepared(
            prepared=prepared,
            batch_size=batch_size,
            delay_seconds=delay_seconds,
            language=language,
        )
        rows = [row.model_dump() for row in result.rows]
        if file_format == "csv":
            return self._export_csv(rows)
        if file_format != "xlsx":
            raise ValueError("file_format must be csv or xlsx")
        return self._export_xlsx(rows)

    def export_rows(self, rows: list[dict[str, Any]], file_format: str = "xlsx") -> tuple[str, bytes, str]:
        if file_format == "csv":
            return self._export_csv(rows)
        if file_format != "xlsx":
            raise ValueError("file_format must be csv or xlsx")
        return self._export_xlsx(rows)

    def prepare_content(
        self,
        file_name: str,
        content: bytes,
        mapping_override: dict[str, str | None] | None = None,
        sheet_name: str | None = None,
        validate_required_mapping: bool = True,
    ) -> dict[str, Any]:
        if not content:
            raise ValueError("Uploaded file is empty")

        suffix = self._suffix(file_name)
        if suffix not in {".csv", ".xlsx", ".xls"}:
            raise ValueError("Unsupported file type")

        selected_sheet, data_frame = self._read_dataframe(content, suffix, sheet_name)
        rows = self._serialize_rows(data_frame)
        columns = list(data_frame.columns)
        mapping = self._detect_mapping(columns)
        if mapping_override:
            mapping.update(mapping_override)
        if validate_required_mapping and not mapping.get("order_number"):
            raise ValueError("Could not identify an order number column")

        deduped_rows: list[LiteInputRow] = []
        preview_rows: list[dict[str, str | None]] = []
        seen_pairs: set[tuple[str, str | None]] = set()
        missing_order_rows = 0
        duplicate_pairs_removed = 0
        query_tracking_numbers: list[str] = []

        # 업로드 원본은 주문 상세 행 기준이므로
        # 주문번호 없는 행 제거 -> (주문번호, 송장번호) 기준 중복 제거 -> 조회용 송장만 unique 추출
        # 순서로 정리한다.
        for row in rows:
            order_number = self._extract_field(row, mapping, "order_number")
            tracking_number = self._normalize_tracking_number(self._extract_field(row, mapping, "tracking_number"))
            if not order_number:
                missing_order_rows += 1
                continue

            pair_key = (order_number, tracking_number)
            if pair_key in seen_pairs:
                duplicate_pairs_removed += 1
                continue
            seen_pairs.add(pair_key)

            item = LiteInputRow(order_number=order_number, tracking_number=tracking_number)
            deduped_rows.append(item)
            if len(preview_rows) < 20:
                preview_rows.append({"order_number": order_number, "tracking_number": tracking_number})
            if tracking_number:
                query_tracking_numbers.append(tracking_number)

        query_tracking_numbers = list(dict.fromkeys(query_tracking_numbers))
        no_tracking_rows = sum(1 for item in deduped_rows if not item.tracking_number)

        return {
            "file_name": file_name,
            "selected_sheet": selected_sheet,
            "columns": columns,
            "mapping": mapping,
            "total_rows": len(rows),
            "rows": deduped_rows,
            "preview_rows": preview_rows,
            "missing_order_rows": missing_order_rows,
            "duplicate_pairs_removed": duplicate_pairs_removed,
            "query_target_count": len(query_tracking_numbers),
            "query_tracking_numbers": query_tracking_numbers,
            "no_tracking_rows": no_tracking_rows,
        }

    def run_prepared(
        self,
        prepared: dict[str, Any],
        batch_size: int,
        delay_seconds: float,
        language: str,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> LiteRunResponse:
        route_map = self._fetch_route_map(
            tracking_numbers=prepared["query_tracking_numbers"],
            batch_size=batch_size,
            delay_seconds=delay_seconds,
            language=language,
            progress_callback=progress_callback,
        )
        return self._build_run_response(prepared, route_map)

    def _build_analyze_response(self, prepared: dict[str, Any]) -> LiteAnalyzeResponse:
        return LiteAnalyzeResponse(
            file_name=prepared["file_name"],
            selected_sheet=prepared["selected_sheet"],
            columns=prepared["columns"],
            detected_mapping=prepared["mapping"],
            total_rows=prepared["total_rows"],
            missing_order_rows=prepared["missing_order_rows"],
            duplicate_pairs_removed=prepared["duplicate_pairs_removed"],
            deduped_rows=len(prepared["rows"]),
            query_target_count=prepared["query_target_count"],
            no_tracking_rows=prepared["no_tracking_rows"],
            preview_rows=[LitePreviewRow(**row) for row in prepared["preview_rows"]],
        )

    def _build_run_response(
        self,
        prepared: dict[str, Any],
        route_map: dict[str, LiteStatusResult],
    ) -> LiteRunResponse:
        rows: list[LiteResultRow] = []
        status_counts: Counter[str] = Counter()
        for item in prepared["rows"]:
            if not item.tracking_number:
                result = LiteResultRow(
                    order_number=item.order_number,
                    tracking_number=None,
                    status="NO_TRACKING",
                )
            else:
                mapped = route_map.get(item.tracking_number)
                if mapped is None:
                    result = LiteResultRow(
                        order_number=item.order_number,
                        tracking_number=item.tracking_number,
                        status="QUERY_FAILED",
                        sf_express_remark="No route response returned by SF",
                    )
                else:
                    result = LiteResultRow(
                        order_number=item.order_number,
                        tracking_number=item.tracking_number,
                        status=mapped.status,
                        sf_express_code=mapped.sf_express_code,
                        sf_express_remark=mapped.sf_express_remark,
                        last_event_time=mapped.last_event_time,
                        latest_event=mapped.latest_event,
                    )
            rows.append(result)
            status_counts[result.status] += 1

        summary = LiteRunSummary(
            total_rows=prepared["total_rows"],
            missing_order_rows=prepared["missing_order_rows"],
            duplicate_pairs_removed=prepared["duplicate_pairs_removed"],
            deduped_rows=len(prepared["rows"]),
            query_target_count=prepared["query_target_count"],
            no_tracking_rows=prepared["no_tracking_rows"],
            status_counts=dict(status_counts),
        )

        return LiteRunResponse(
            file_name=prepared["file_name"],
            selected_sheet=prepared["selected_sheet"],
            detected_mapping=prepared["mapping"],
            summary=summary,
            rows=rows,
        )

    async def _read_upload(self, upload_file: UploadFile) -> tuple[str, bytes]:
        file_name = upload_file.filename or "uploaded-file"
        content = await upload_file.read()
        return file_name, content

    def _read_dataframe(self, content: bytes, suffix: str, sheet_name: str | None) -> tuple[str | None, pd.DataFrame]:
        if suffix == ".csv":
            data_frame = pd.read_csv(BytesIO(content), dtype=str, keep_default_na=False)
            return None, self._normalize_dataframe(data_frame)

        excel_file = pd.ExcelFile(BytesIO(content))
        selected_sheet = sheet_name
        if selected_sheet is None:
            selected_sheet = self._select_sheet(excel_file)
        data_frame = pd.read_excel(BytesIO(content), sheet_name=selected_sheet, dtype=str, keep_default_na=False)
        return selected_sheet, self._normalize_dataframe(data_frame)

    def _select_sheet(self, excel_file: pd.ExcelFile) -> str:
        best_sheet = excel_file.sheet_names[0]
        best_score = -1
        for candidate in excel_file.sheet_names:
            frame = self._normalize_dataframe(
                pd.read_excel(excel_file, sheet_name=candidate, dtype=str, keep_default_na=False)
            )
            mapping = self._detect_mapping(list(frame.columns))
            score = frame.shape[0]
            if mapping.get("tracking_number"):
                score += 100
            if mapping.get("order_number"):
                score += 1000
            if score > best_score:
                best_sheet = candidate
                best_score = score
        return best_sheet

    def _normalize_dataframe(self, data_frame: pd.DataFrame) -> pd.DataFrame:
        normalized = data_frame.copy()
        normalized.columns = [str(column).strip() for column in normalized.columns]
        for column in normalized.columns:
            normalized[column] = normalized[column].astype(str).str.strip()
            normalized.loc[normalized[column].isin(["", "nan", "None"]), column] = ""
        return normalized

    def _serialize_rows(self, data_frame: pd.DataFrame) -> list[dict[str, Any]]:
        rows = []
        for row in data_frame.to_dict(orient="records"):
            rows.append({str(key): (None if value == "" else value) for key, value in row.items()})
        return rows

    def _detect_mapping(self, columns: list[str]) -> dict[str, str | None]:
        normalized = {self._normalize(column): column for column in columns}
        mapping: dict[str, str | None] = {}
        for target_field, aliases in FIELD_ALIASES.items():
            mapping[target_field] = None
            for alias in aliases:
                column = normalized.get(self._normalize(alias))
                if column:
                    mapping[target_field] = column
                    break
        return mapping

    def _extract_field(self, row: dict[str, Any], mapping: dict[str, str | None], field_name: str) -> str | None:
        column_name = mapping.get(field_name)
        if not column_name:
            return None
        value = row.get(column_name)
        if value is None:
            return None
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return None
        return text

    def _fetch_route_map(
        self,
        tracking_numbers: list[str],
        batch_size: int,
        delay_seconds: float,
        language: str,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> dict[str, LiteStatusResult]:
        if not tracking_numbers:
            if progress_callback:
                progress_callback(0, 0)
            return {}
        batch_size = max(1, min(batch_size, 10))
        active_api_key = self.settings_service.get_active_api_key()
        if active_api_key is None:
            raise ValueError("No active SF API key configured")

        api_key_record, secret_fields = active_api_key
        credentials = SFClientCredentials(
            partner_id=secret_fields["partner_id"],
            checkword=secret_fields["checkword"],
            environment=api_key_record.environment,
        )

        total_targets = len(tracking_numbers)
        completed_targets = 0
        if progress_callback:
            progress_callback(completed_targets, total_targets)

        route_map: dict[str, LiteStatusResult] = {}
        concurrency = max(1, min(self.settings.lite_fetch_concurrency, MAX_LITE_FETCH_CONCURRENCY))
        batches = [tracking_numbers[start : start + batch_size] for start in range(0, len(tracking_numbers), batch_size)]

        # SF는 배치 10건까지가 안정적이어서 10건 단위로 끊고,
        # 한 번의 run 안에서는 같은 HTTP Client를 재사용한다.
        with SFClient.create_http_client(self.settings) as http_client:
            client = SFClient(
                self.settings,
                credentials,
                http_client=http_client,
            )

            with ThreadPoolExecutor(max_workers=concurrency) as executor:
                futures: dict[Future[dict[str, LiteStatusResult]], list[str]] = {}
                for index, batch in enumerate(batches):
                    futures[executor.submit(self._fetch_route_batch, client, batch, language)] = batch
                    if delay_seconds > 0 and index + 1 < len(batches):
                        time.sleep(delay_seconds)

                for future in as_completed(futures):
                    batch = futures[future]
                    batch_result = future.result()
                    route_map.update(batch_result)
                    # 병렬 실행이어도 "배치 완료 수" 기준으로만 진행률을 올려야 역행이 생기지 않는다.
                    completed_targets += len(batch)
                    if progress_callback:
                        progress_callback(completed_targets, total_targets)
        return route_map

    def _fetch_route_batch(
        self,
        client: SFClient,
        batch: list[str],
        language: str,
    ) -> dict[str, LiteStatusResult]:
        try:
            response = client.search_routes(batch, language=language)
            route_resps, _ = client.extract_route_payload(response)
            batch_map = {
                str(item.get("mailNo") or item.get("trackingNumber") or "").strip(): item
                for item in route_resps
                if str(item.get("mailNo") or item.get("trackingNumber") or "").strip()
            }
            results: dict[str, LiteStatusResult] = {}
            for tracking_number in batch:
                route_resp = batch_map.get(tracking_number)
                if route_resp is None:
                    # SF가 배치 응답에서 특정 송장을 빠뜨리는 경우가 있어,
                    # 이런 건 NO_ROUTE가 아니라 조회 실패로 남긴다.
                    results[tracking_number] = LiteStatusResult(
                        status="QUERY_FAILED",
                        sf_express_code=PARTIAL_MISSING_CODE,
                        sf_express_remark=f"{PARTIAL_MISSING_REMARK}: {tracking_number}",
                        last_event_time=None,
                        latest_event=None,
                    )
                    continue
                results[tracking_number] = map_route_response(route_resp)
            return results
        except SFClientError as error:
            return {
                tracking_number: LiteStatusResult(
                    status="QUERY_FAILED",
                    sf_express_code="SF_CLIENT_ERROR",
                    sf_express_remark=str(error),
                    last_event_time=None,
                    latest_event=None,
                )
                for tracking_number in batch
            }
        except Exception as error:  # pragma: no cover - defensive local workflow
            return {
                tracking_number: LiteStatusResult(
                    status="QUERY_FAILED",
                    sf_express_code="UNEXPECTED_ERROR",
                    sf_express_remark=str(error),
                    last_event_time=None,
                    latest_event=None,
                )
                for tracking_number in batch
            }

    def _export_xlsx(self, rows: list[dict[str, Any]]) -> tuple[str, bytes, str]:
        workbook = Workbook()
        # 운영자가 가장 먼저 보는 상태를 분리하려고 도착/수취완료 시트를 따로 둔다.
        primary_rows = [row for row in rows if row.get("status") in {"ARRIVED", "COLLECTED"}]
        secondary_rows = [row for row in rows if row.get("status") not in {"ARRIVED", "COLLECTED"}]
        unknown_rows = [row for row in rows if row.get("status") == "UNKNOWN"]

        worksheet = workbook.active
        worksheet.title = "ARRIVED_COLLECTED"
        self._append_export_rows(worksheet, primary_rows)

        other_sheet = workbook.create_sheet(title="OTHER_STATUS")
        self._append_export_rows(other_sheet, secondary_rows)

        if unknown_rows:
            # UNKNOWN은 나중에 매핑 규칙을 보강해야 하므로 최신 이벤트 원문까지 별도 시트로 남긴다.
            unknown_sheet = workbook.create_sheet(title="UNKNOWN_LOG")
            self._append_export_rows(
                unknown_sheet,
                self._build_unknown_log_rows(unknown_rows),
                columns=UNKNOWN_LOG_COLUMNS,
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return (
            "lite-tracking-results.xlsx",
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _export_csv(self, rows: list[dict[str, Any]]) -> tuple[str, bytes, str]:
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=[header for header, _ in EXPORT_COLUMNS])
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    header: escape_excel_formula(self._export_value(field, row.get(field)))
                    for header, field in EXPORT_COLUMNS
                }
            )
        return ("lite-tracking-results.csv", buffer.getvalue().encode("utf-8-sig"), "text/csv")

    def _append_export_rows(
        self,
        worksheet: Any,
        rows: list[dict[str, Any]],
        columns: list[tuple[str, str]] = EXPORT_COLUMNS,
    ) -> None:
        worksheet.append([header for header, _ in columns])
        for row in rows:
            worksheet.append([escape_excel_formula(self._export_value(field, row.get(field))) for _, field in columns])
        self._style_export_sheet(worksheet)

    def _export_value(self, field: str, value: Any) -> Any:
        if field == "status":
            return EXPORT_STATUS_LABELS.get(str(value), value)
        return value

    def _build_unknown_log_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        log_rows: list[dict[str, Any]] = []
        for row in rows:
            latest_event = row.get("latest_event") or {}
            # UNKNOWN 원인은 코드 조합만 보면 부족할 수 있어서 최신 이벤트 원문을 같이 저장한다.
            log_rows.append(
                {
                    "order_number": row.get("order_number"),
                    "tracking_number": row.get("tracking_number"),
                    "status": row.get("status"),
                    "sf_express_code": row.get("sf_express_code"),
                    "sf_express_remark": row.get("sf_express_remark"),
                    "last_event_time": row.get("last_event_time"),
                    "op_code": latest_event.get("opCode") or latest_event.get("opcode"),
                    "first_status_code": latest_event.get("firstStatusCode") or latest_event.get("first_status_code"),
                    "secondary_status_code": (
                        latest_event.get("secondaryStatusCode")
                        or latest_event.get("secondStatusCode")
                        or latest_event.get("secondary_status_code")
                    ),
                    "latest_event_raw": json.dumps(latest_event, ensure_ascii=False, sort_keys=True)
                    if latest_event
                    else None,
                }
            )
        return log_rows

    def _style_export_sheet(self, worksheet: Any) -> None:
        for cell in worksheet[1]:
            cell.fill = EXPORT_HEADER_FILL
            cell.font = EXPORT_HEADER_FONT
            cell.alignment = EXPORT_HEADER_ALIGNMENT

        # 엑셀은 기본 폭이 좁아서 한글/중문 remark가 잘리는 경우가 많아
        # 실제 표시 폭 기준으로 컬럼 너비를 계산한다.
        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            max_width = 0
            for row_index in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                display_width = self._display_text_width(value)
                if display_width > max_width:
                    max_width = display_width
            worksheet.column_dimensions[column_letter].width = max(12, max_width + 2)

    def _display_text_width(self, value: Any) -> int:
        if value is None:
            return 0

        text = str(value)
        width = 0
        for character in text:
            width += 2 if unicodedata.east_asian_width(character) in {"F", "W", "A"} else 1
        return width

    def _normalize(self, value: str) -> str:
        return "".join(character.lower() for character in value if character.isalnum())

    def _normalize_tracking_number(self, value: str | None) -> str | None:
        if not value:
            return None
        normalized = "".join(value.split()).upper()
        return normalized or None

    def _suffix(self, filename: str) -> str:
        lowered = filename.lower()
        if lowered.endswith(".xlsx"):
            return ".xlsx"
        if lowered.endswith(".xls"):
            return ".xls"
        if lowered.endswith(".csv"):
            return ".csv"
        return ""
