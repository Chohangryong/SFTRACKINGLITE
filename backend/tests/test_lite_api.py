from __future__ import annotations

import json
import time
from io import BytesIO

from openpyxl import load_workbook


def test_lite_analyze_dedupes_pairs_and_counts_missing_tracking(client) -> None:
    response = client.post(
        "/api/lite/analyze",
        files={
            "file": (
                "orders.csv",
                (
                    "Order Number,Tracking Number\n"
                    "ORDER-1,SF111\n"
                    "ORDER-1,SF111\n"
                    "ORDER-1,SF222\n"
                    ",SF333\n"
                    "ORDER-2,\n"
                ).encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["detected_mapping"] == {
        "order_number": "Order Number",
        "tracking_number": "Tracking Number",
    }
    assert payload["total_rows"] == 5
    assert payload["missing_order_rows"] == 1
    assert payload["duplicate_pairs_removed"] == 1
    assert payload["deduped_rows"] == 3
    assert payload["query_target_count"] == 2
    assert payload["no_tracking_rows"] == 1


def test_lite_run_uses_analyzed_status_rules(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )

    mocked_response = {
        "apiResultCode": "A1000",
        "apiResultData": json.dumps(
            {
                "errorCode": "S0000",
                "routeResps": [
                    {
                        "mailNo": "SFARRIVED",
                        "routes": [
                            {
                                "acceptTime": "2026-03-10 09:00:00",
                                "opCode": "80",
                                "firstStatusCode": "4",
                                "secondaryStatusCode": "401",
                                "remark": "\u5df2\u6d3e\u9001\u81f3\uff08\u987a\u4e30\u81ea\u52a9\u67dc\uff09",
                            }
                        ],
                    },
                    {
                        "mailNo": "SFCOLLECTED",
                        "routes": [
                            {
                                "acceptTime": "2026-03-10 10:00:00",
                                "opCode": "80",
                                "firstStatusCode": "4",
                                "secondaryStatusCode": "401",
                                "remark": "\u60a8\u7684\u5feb\u4ef6\u5df2\u6d3e\u9001\u81f3\u672c\u4eba",
                            }
                        ],
                    },
                    {
                        "mailNo": "SFSHIPPED",
                        "routes": [
                            {
                                "acceptTime": "2026-03-10 11:00:00",
                                "opCode": "634",
                                "firstStatusCode": "3",
                                "secondaryStatusCode": "301",
                                "remark": "\u5feb\u4ef6\u6b63\u5728\u6d3e\u9001\u9014\u4e2d",
                            }
                        ],
                    },
                    {
                        "mailNo": "SFUNAVAILABLE",
                        "routes": [],
                        "reasonCode": "20028",
                        "reasonRemark": "\u6708\u7ed3\u5361\u53f7\u4e0d\u5339\u914d",
                    },
                    {
                        "mailNo": "SFNOROUTE",
                        "routes": [],
                    },
                ],
            }
        ),
    }
    mocker.patch("app.services.sf_client.SFClient.search_routes", return_value=mocked_response)

    response = client.post(
        "/api/lite/run",
        files={
            "file": (
                "orders.csv",
                (
                    "Order Number,Tracking Number\n"
                    "ORDER-1,SFARRIVED\n"
                    "ORDER-2,SFCOLLECTED\n"
                    "ORDER-3,SFSHIPPED\n"
                    "ORDER-4,SFUNAVAILABLE\n"
                    "ORDER-5,SFNOROUTE\n"
                    "ORDER-6,\n"
                ).encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    rows = {row["order_number"]: row for row in payload["rows"]}
    assert rows["ORDER-1"]["status"] == "ARRIVED"
    assert rows["ORDER-2"]["status"] == "COLLECTED"
    assert rows["ORDER-3"]["status"] == "SHIPPED"
    assert rows["ORDER-4"]["status"] == "QUERY_UNAVAILABLE"
    assert rows["ORDER-4"]["sf_express_code"] == "20028"
    assert rows["ORDER-5"]["status"] == "NO_ROUTE"
    assert rows["ORDER-6"]["status"] == "NO_TRACKING"
    assert payload["summary"]["status_counts"] == {
        "ARRIVED": 1,
        "COLLECTED": 1,
        "SHIPPED": 1,
        "QUERY_UNAVAILABLE": 1,
        "NO_ROUTE": 1,
        "NO_TRACKING": 1,
    }


def test_lite_export_uses_customer_headers(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )
    mocker.patch(
        "app.services.sf_client.SFClient.search_routes",
        return_value={
            "apiResultCode": "A1000",
            "apiResultData": json.dumps(
                {
                    "errorCode": "S0000",
                    "routeResps": [
                        {
                            "mailNo": "SF123",
                            "routes": [
                                {
                                    "acceptTime": "2026-03-10 10:00:00",
                                    "opCode": "80",
                                    "firstStatusCode": "4",
                                    "secondaryStatusCode": "401",
                                    "remark": "\u5df2\u6d3e\u9001\u81f3\uff08\u987a\u4e30\u81ea\u52a9\u67dc\uff09",
                                }
                            ],
                        },
                        {
                            "mailNo": "SF456",
                            "routes": [
                                {
                                    "acceptTime": "2026-03-10 10:30:00",
                                    "opCode": "634",
                                    "firstStatusCode": "3",
                                    "secondaryStatusCode": "301",
                                    "remark": "\u5feb\u4ef6\u6b63\u5728\u6d3e\u9001\u9014\u4e2d",
                                }
                            ],
                        },
                    ],
                }
            ),
        },
    )

    response = client.post(
        "/api/lite/export",
        data={"file_format": "xlsx"},
        files={
            "file": (
                "orders.csv",
                "Order Number,Tracking Number\nORDER-1,SF123\nORDER-2,SF456\n".encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["ARRIVED_COLLECTED", "OTHER_STATUS"]

    primary_sheet = workbook["ARRIVED_COLLECTED"]
    secondary_sheet = workbook["OTHER_STATUS"]
    headers = [cell.value for cell in primary_sheet[1]]
    assert headers == [
        "쇼핑몰오더번호",
        "송장번호",
        "송장상태",
        "택배사최신상태코드",
        "택배사최신REMARK",
    ]
    assert primary_sheet["A1"].font.bold is True
    assert primary_sheet["A1"].fill.fill_type == "solid"
    assert secondary_sheet["A1"].font.bold is True
    assert secondary_sheet["A1"].fill.fill_type == "solid"
    assert primary_sheet.column_dimensions["E"].width > len("택배사최신REMARK")
    assert primary_sheet["A2"].value == "ORDER-1"
    assert primary_sheet["C2"].value == "ARRIVED"
    assert secondary_sheet["A2"].value == "ORDER-2"
    assert secondary_sheet["C2"].value == "SHIPPED"


def test_lite_export_localizes_query_unavailable_status(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )
    mocker.patch(
        "app.services.sf_client.SFClient.search_routes",
        return_value={
            "apiResultCode": "A1000",
            "apiResultData": json.dumps(
                {
                    "errorCode": "S0000",
                    "routeResps": [
                        {
                            "mailNo": "SFUNAVAILABLE",
                            "routes": [],
                            "reasonCode": "20028",
                            "reasonRemark": "\u6708\u7ed3\u5361\u53f7\u4e0d\u5339\u914d",
                        }
                    ],
                }
            ),
        },
    )

    response = client.post(
        "/api/lite/export",
        data={"file_format": "xlsx"},
        files={
            "file": (
                "orders.csv",
                "Order Number,Tracking Number\nORDER-1,SFUNAVAILABLE\n".encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    secondary_sheet = workbook["OTHER_STATUS"]
    assert secondary_sheet["C2"].value == "\uC870\uD68C\uBD88\uAC00"


def test_lite_export_adds_unknown_log_sheet(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )
    mocker.patch(
        "app.services.sf_client.SFClient.search_routes",
        return_value={
            "apiResultCode": "A1000",
            "apiResultData": json.dumps(
                {
                    "errorCode": "S0000",
                    "routeResps": [
                        {
                            "mailNo": "SFUNKNOWN",
                            "routes": [
                                {
                                    "acceptTime": "2026-03-10 10:45:00",
                                    "opCode": "999",
                                    "firstStatusCode": "9",
                                    "secondaryStatusCode": "909",
                                    "remark": "Unmapped state",
                                }
                            ],
                        }
                    ],
                }
            ),
        },
    )

    response = client.post(
        "/api/lite/export",
        data={"file_format": "xlsx"},
        files={
            "file": (
                "orders.csv",
                "Order Number,Tracking Number\nORDER-1,SFUNKNOWN\n".encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["ARRIVED_COLLECTED", "OTHER_STATUS", "UNKNOWN_LOG"]

    other_sheet = workbook["OTHER_STATUS"]
    unknown_sheet = workbook["UNKNOWN_LOG"]

    assert other_sheet["C2"].value == "UNKNOWN"
    assert [cell.value for cell in unknown_sheet[1]] == [
        "쇼핑몰오더번호",
        "송장번호",
        "송장상태",
        "택배사최신상태코드",
        "택배사최신REMARK",
        "최신이벤트일시",
        "OP CODE",
        "1차상태코드",
        "2차상태코드",
        "최신이벤트원문",
    ]
    assert unknown_sheet["A2"].value == "ORDER-1"
    assert unknown_sheet["C2"].value == "UNKNOWN"
    assert unknown_sheet["G2"].value == "999"
    assert unknown_sheet["H2"].value == "9"
    assert unknown_sheet["I2"].value == "909"
    assert '"opCode": "999"' in unknown_sheet["J2"].value


def test_lite_run_marks_partial_batch_missing_as_query_failed(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )

    mocker.patch(
        "app.services.sf_client.SFClient.search_routes",
        return_value={
            "apiResultCode": "A1000",
            "apiResultData": json.dumps(
                {
                    "errorCode": "S0000",
                    "routeResps": [
                        {
                            "mailNo": "SFOK",
                            "routes": [
                                {
                                    "acceptTime": "2026-03-10 10:00:00",
                                    "opCode": "634",
                                    "firstStatusCode": "3",
                                    "secondaryStatusCode": "301",
                                    "remark": "\u5feb\u4ef6\u6b63\u5728\u6d3e\u9001\u9014\u4e2d",
                                }
                            ],
                        }
                    ],
                }
            ),
        },
    )

    response = client.post(
        "/api/lite/run",
        files={
            "file": (
                "orders.csv",
                "Order Number,Tracking Number\nORDER-1,SFOK\nORDER-2,SFMISSING\n".encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    rows = {row["order_number"]: row for row in response.json()["rows"]}
    assert rows["ORDER-1"]["status"] == "SHIPPED"
    assert rows["ORDER-2"]["status"] == "QUERY_FAILED"
    assert rows["ORDER-2"]["sf_express_code"] == "SF_PARTIAL_MISSING"
    assert "SFMISSING" in rows["ORDER-2"]["sf_express_remark"]


def test_lite_job_progress_is_monotonic_under_parallel_batches(client, mocker) -> None:
    client.post(
        "/api/settings/api-keys",
        json={
            "label": "Production",
            "environment": "production",
            "partner_id": "PARTNER",
            "checkword": "CHECKWORD",
            "is_active": True,
        },
    )

    def mocked_search_routes(_self, tracking_numbers, language=None):
        time.sleep(0.05)
        return {
            "apiResultCode": "A1000",
            "apiResultData": json.dumps(
                {
                    "errorCode": "S0000",
                    "routeResps": [
                        {
                            "mailNo": tracking_number,
                            "routes": [
                                {
                                    "acceptTime": "2026-03-10 10:00:00",
                                    "opCode": "634",
                                    "firstStatusCode": "3",
                                    "secondaryStatusCode": "301",
                                    "remark": "\u5feb\u4ef6\u6b63\u5728\u6d3e\u9001\u9014\u4e2d",
                                }
                            ],
                        }
                        for tracking_number in tracking_numbers
                    ],
                }
            ),
        }

    mocker.patch("app.services.sf_client.SFClient.search_routes", autospec=True, side_effect=mocked_search_routes)

    csv_rows = ["Order Number,Tracking Number"]
    for index in range(1, 22):
        csv_rows.append(f"ORDER-{index},SF{index:03d}")

    create_response = client.post(
        "/api/lite/jobs",
        files={
            "file": (
                "orders.csv",
                ("\n".join(csv_rows) + "\n").encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert create_response.status_code == 200
    job_id = create_response.json()["job_id"]

    progress_values: list[int] = []
    completed_values: list[int] = []
    status = "queued"
    for _ in range(30):
        poll_response = client.get(f"/api/lite/jobs/{job_id}")
        assert poll_response.status_code == 200
        payload = poll_response.json()
        progress_values.append(payload["progress_percent"])
        completed_values.append(payload["completed_targets"])
        status = payload["status"]
        if status == "completed":
            break
        time.sleep(0.03)

    assert status == "completed"
    assert progress_values == sorted(progress_values)
    assert completed_values == sorted(completed_values)
    assert completed_values[-1] == 21
    assert progress_values[-1] == 100
